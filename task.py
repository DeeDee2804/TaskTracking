import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border
from datetime import datetime
from typing import Dict, List, Optional


COLUMN_MAPPING = {
    "Do Date": "do_date",
    "Category task": "category",
    "Task": "task",
    "Description": "description",
    "Assigning person": "assigner",
    "Deadline": "deadline",
    "Status": "status",
    "Estimated (h)": "estimated_hours",
    "Spent (h)": "spent_hours",
    "Result": "result",
    "Reason": "reason"
}
INTERNAL_COLUMN = [
     "do_date",
     "category",
     "task",
     "description",
     "assigner",
     "deadline",
     "status",
     "estimated_hours",
     "spent_hours",
     "result",
     "reason"              
]

def swap_key_dict(input_dict: Dict[str, str]) -> Dict[str, str]:
    """
    Swap the keys and values in the provided dictionary.

    Args:
        input_dict (Dict[str, str]): Dictionary with keys and values to swap.

    Returns:
        Dict[str, str]: New dictionary with keys and values swapped.
    """
    return {value: key for key, value in input_dict.items()}

CONVERTED_COLUMN = swap_key_dict(COLUMN_MAPPING)

def load_task_list(path) -> List[Dict[str, Optional[str]]]:
    """
    Load tasks from the Excel database and return as a list of dictionaries.

    Returns:
        List[Dict[str, Optional[str]]]: List of task dictionaries with formatted dates and NaN values replaced.
    """
    try:
        data = pd.read_excel(path, usecols=COLUMN_MAPPING.keys())
    except Exception as e:
        raise RuntimeError(f"Failed to load data from {path}: {e}")

    data.rename(columns=COLUMN_MAPPING, inplace=True)

    task_list = []
    for _, row in data.iterrows():
        task_item = row.to_dict()
        # Format dates if present
        for date_field in ['do_date', 'deadline']:
            if type(task_item[date_field]) == datetime or type(task_item[date_field]) == pd.Timestamp:
                task_item[date_field] = task_item[date_field].strftime('%Y-%m-%d')

        # Replace NaN values with empty strings
        task_item = {key: ("" if (pd.isna(value)) else value)
                     for key, value in task_item.items()}

        task_list.append(task_item)

    return task_list

def edit_task_item(path, index: int, data: Dict[str, Optional[str]]):
    """
    Edit an existing task item in the Excel database.

    Args:
        index (int): Index of the row to edit (1-based).
        data (Dict[str, Optional[str]]): Dictionary of data to update the task item with.
    """
    try:
        wb = load_workbook(path)
        ws = wb.active

        for col_num in range(1, len(data)+1):
            # Adjust row index for 1-based and not count the header row
            ws.cell(row=index + 2, column=col_num, value=data[INTERNAL_COLUMN[col_num-1]])  
        wb.save(path)
    except Exception as e:
        raise RuntimeError(f"Failed to edit task item: {e}")

def delete_task_item(path:str, index: int):
    """
    Edit an existing task item in the Excel database.

    Args:
        index (int): Index of the row to edit (1-based).
        data (Dict[str, Optional[str]]): Dictionary of data to update the task item with.
    """
    try:
        wb = load_workbook(path)
        ws = wb.active
        
        # Adjust row index for 1-based and not count the header row
        ws.delete_rows(index + 2)
        
        wb.save(path)
    except Exception as e:
        raise RuntimeError(f"Failed to edit task item: {e}")
    
def add_new_task_item(path, data: Dict[str, Optional[str]]):
    """
    Add a new task item to the Excel database.

    Args:
        data (Dict[str, Optional[str]]): Dictionary of data to add as a new task item.
    """
    try:
        wb = load_workbook(path)
        ws = wb.active

        last_row = ws.max_row
        for col_num in range(1, len(data)+1):
            ws.cell(row=last_row + 1, column=col_num, value=data[INTERNAL_COLUMN[col_num-1]])

        # Copy formatting from the previous row to the new row
        for col_num in range(1, ws.max_column + 1):
            prev_cell = ws.cell(row=last_row, column=col_num)
            new_cell = ws.cell(row=last_row + 1, column=col_num)

            # Copy cell styles
            if prev_cell.font:
                new_cell.font = Font(
                    name=prev_cell.font.name,
                    bold=prev_cell.font.bold,
                    italic=prev_cell.font.italic,
                    vertAlign=prev_cell.font.vertAlign,
                    underline=prev_cell.font.underline,
                    strike=prev_cell.font.strike,
                    color=prev_cell.font.color
                )

            if prev_cell.fill:
                new_cell.fill = PatternFill(
                    fill_type=prev_cell.fill.fill_type,
                    start_color=prev_cell.fill.start_color,
                    end_color=prev_cell.fill.end_color
                )

            if prev_cell.border:
                new_cell.border = Border(
                    left=prev_cell.border.left,
                    right=prev_cell.border.right,
                    top=prev_cell.border.top,
                    bottom=prev_cell.border.bottom
                )

            if prev_cell.alignment:
                new_cell.alignment = Alignment(
                    horizontal=prev_cell.alignment.horizontal,
                    vertical=prev_cell.alignment.vertical,
                    text_rotation=prev_cell.alignment.text_rotation,
                    wrap_text=prev_cell.alignment.wrap_text,
                    shrink_to_fit=prev_cell.alignment.shrink_to_fit,
                    indent=prev_cell.alignment.indent
                )

            if prev_cell.number_format:
                new_cell.number_format = prev_cell.number_format

        wb.save(path)
    except Exception as e:
        raise RuntimeError(f"Failed to add new task item: {e}")
